from flask import Flask, render_template, request
from openpyxl import load_workbook

wb = load_workbook("Colloscope PCSI 2025-2026 S1&S2.xlsx")
print(wb)

app = Flask(__name__)

#1ere page avec le formulaire

@app.route('/')
def welcome():
    return render_template("welcome.html")

#2eme page avec les horaires

@app.route("/calendrier", methods=["POST"])
def calendrier():
    amail = request.form["amail"]
    semaine = int(request.form["semaine"])
    #recuperer les horaires dans le xlsx :
    #trouver le groupe
    sheet = wb["Créneaux&Salles&listesGroupes"]
    groupe = None
    for row in sheet.iter_rows(values_only=True):
        if amail==row[10]:
            groupe = row[7]
            nom = row[8]
            prenom = row[9]
            break
    
    #trouver le code colle (ex : A5, P3...)
    sheet = wb["Colloscope pour les étudiants"]
    #code_colle = sheet.cell(row=semaine+3, column=groupe+1).value
    code_colle = None
    for row in sheet.iter_rows(values_only=True):
        if semaine==row[1]:
            code_colle = row[int(groupe+1)]
            break

    #trouver le code horaire (ex : code_horaire1="A5")
    code_horaire1 = code_colle.split(",")[0].strip()
    code_horaire2 = code_colle.split(",")[1].strip()

    #trouver les informations des horaires (ex : info_horaire1=....)
    sheet = wb["Créneaux&Salles&listesGroupes"]
    for row in sheet.iter_rows(values_only=True):
        if code_horaire1==row[1]:
            i_horaire1 = [cell for cell in row[1:5]]
            break
    
    for row in sheet.iter_rows(values_only=True):
        if code_horaire2==row[1]:
            i_horaire2 = [cell for cell in row[1:5]]
            break


    #pas toucher
    if groupe==None:
        return render_template("introuvable.html")
    else:
        return render_template("calendrier.html", groupe=groupe, prenom=prenom, nom=nom, horaire1=i_horaire1, horaire2=i_horaire2)



if __name__ == "__main__":
    app.run(debug=True)